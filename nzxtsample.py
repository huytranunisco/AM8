import time
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import BNPauto

start_time = time.time()

billingAcc = 'NZXT'
accName = 'NZXT'
facility = 'Valley View'
startPeriod = '08/01/22'
endPeriod = '08/15/22'
user = 'kevin'
billCycle = 'Bimonthly'

facility = BNPauto.exportHandle(billingAcc, facility, startPeriod, endPeriod, user)
facility = facility.lower()
facility = facility.capitalize()

reportLoc = BNPauto.invoiceToReport(user, accName, facility, billCycle)

#Input file path for activity report
activityLoc = r"C:\\Users\\kevin\\Documents\\AUTOM8\\NZXT\\activityReoport(NZXT - 2022-08-01 - 2022-08-15).xlsx"

report = pd.read_excel(reportLoc)
billingitems = report['Description'].tolist()
for count, name in enumerate(billingitems):
    billingitems[count] = name.lower()

print(billingitems)

book = load_workbook(reportLoc)
writer = pd.ExcelWriter(reportLoc, engine = 'openpyxl')
writer.book = book

def dataCopy(dataframe, sheetName):
    if sheetName in writer.sheets:
        dataframe.to_excel(writer, sheet_name = sheetName,startrow = writer.sheets[sheetName].max_row, index = False, header = False)
    else:
        dataframe.to_excel(writer, sheet_name = sheetName, index = False)

    writer.save()

# accessorial cancel order per order, ispicked,yes; / cancellation charge
def cancelledOrder(arPath):
    df = pd.read_excel(io=arPath, sheet_name='Order & Receipt')
    df = df[df['Shipment'].isin(['OUTBOUND'])]
    df = df[df['STATUS'].isin(['CANCELLED'])]
    count = df['STATUS'].count()

    return df, count

# outbound handling ds : over 750 cartons
def outboundHandlingDS(arPath):
    df = pd.read_excel(io=arPath, sheet_name='Order & Receipt')
    df = df[df['Shipment'].isin(['OUTBOUND'])]
    df = df[df['TYPE'].isin(['DropShip Order'])]
    sum = df['CS QTY'].sum()

    return df, sum

# routing
def routing(arPath):
    df = pd.read_excel(io=arPath, sheet_name='Accessories')
    df = df[df['ACCOUNT ITEM'].isin(['ROUTING'])]
    sum = df['QTY'].sum()

    return df, sum

# receive inbound per carton
def inboundPerCarton(arPath):
    df = pd.read_excel(io=arPath, sheet_name='Order & Receipt')
    df = df[df['Shipment'].isin(['INBOUND'])]
    df = df[df['Offload Type'].isin(['BY_HAND_NO_PALLET'])]
    sum = df['CS QTY'].sum()

    return df, sum

# receive inbound palletized
def inboundPalletized(arPath):
    df = pd.read_excel(io=arPath, sheet_name='Order & Receipt')
    df = df[df['RN/DN'].str.contains('RN', case=False, na=False)]
    df = df[df['Offload Type'].isin(['FORKLIFT_WITH_PALLET'])]
    sum = df['PALLET QTY'].sum()

    return df, sum

# handling pick per pallet, ordertype,rg; / pallet pick (regular order)
def pickPerPalletRG(arPath):
    df = pd.read_excel(io=arPath, sheet_name='Pick Task')
    df = df[df['TYPE'].isin(['Regular Order'])]
    sum = df['PALLETPICKQTY'].sum()

    return df, sum

# storage income initial storage per case / initial storage – per carton
def initialStorage(arPath):
    df = pd.read_excel(io=arPath, sheet_name='Order & Receipt')
    df = df[df['Shipment'].isin(['INBOUND'])]
    sum = df['CS QTY'].sum()

    return df, sum

# handling order processing per order, ordertype,rg;
def handlingOrderProcessingRG(arPath):
    df = pd.read_excel(io=arPath, sheet_name='Order & Receipt')
    df = df[df['STATUS'].isin(['SHIPPED'])]
    df = df[df['TYPE'].isin(['Regular Order'])]
    count = df['TYPE'].count()

    return df, count

# case pick (amazon order) / amazon labeling
def casePickAmazon(arPath):
    df = pd.read_excel(io=arPath, sheet_name='Order & Receipt')
    df = df[df['Shipment'].isin(['OUTBOUND'])]
    df = df[df['TYPE'].isin(['Regular Order', 'DropShip Order'])]
    df = df[df['RETAILER'].isin(['Amazon'])]
    sum = df['CS QTY'].sum()

    return df, sum

# case pick if 30lbs or less (regular order)
def casePick30lbsOrLess(arPath):
    df = pd.read_excel(io=arPath, sheet_name='Pick Task')
    df = df[df['TYPE'].isin(['Regular Order'])]
    df = df[~df['SHIPTO'].isin(['Amazon', 'Amazon.com Services INC.'])]
    sum = df['CASEPICKQTY'].sum() + df['INNERPICKQTY'].sum() + df['PIECEPICKQTY'].sum()

    return df, sum

# rush order
def rushOrder(arPath):
    df = pd.read_excel(io=arPath, sheet_name='Order & Receipt')
    df = df[df['Shipment'].isin(['OUTBOUND'])]
    df = df[df['IS RUSH'].isin(['Y'])]
    count = df['IS RUSH'].count()

    return df, count

for index, item in enumerate(billingitems):
    print(item)
    if item == 'accessorial cancel order per order, ispicked,yes;' or item == 'cancellation charge':
        df, qty = cancelledOrder(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'CANCELLED ORDER')
    elif item == 'outbound handling ds : over 750 cartons':
        df, qty = outboundHandlingDS(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'OUTBOUND HANDLING DS')
    elif item == 'routing':
        df, qty = routing(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'ROUTING')
    elif item == 'receive inbound per carton':
        df, qty = inboundPerCarton(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'INBOUND PER CARTON')
    elif item == 'receive inbound palletized':
        df, qty = inboundPalletized(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'INBOUND PALLETIZED')
    elif item == 'handling pick per pallet, ordertype,rg;' or item == 'pallet pick (regular order)':
        df, qty = pickPerPalletRG(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'PICK PER PALLET RG')
    elif item == 'storage income initial storage per case' or item == 'initial storage – per carton':
        df, qty = initialStorage(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'INITIAL STORAGE')
    elif item == 'handling order processing per order, ordertype,rg;':
        df, qty = handlingOrderProcessingRG(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'HANDLING ORDER PROCESSING RG')
    elif item == 'case pick (amazon order)' or item == 'amazon labeling':
        df, qty = casePickAmazon(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'CASE PICK AMAZON')
    elif item == 'case pick if 30lbs or less (regular order)':
        df, qty = casePick30lbsOrLess(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'CASE PICK 30LBS OR LESS')
    elif item == 'rush order':
        df, qty = rushOrder(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'RUSH ORDER')


report.to_excel(reportLoc, index=False)

print("--- %s seconds ---" % (time.time() - start_time))