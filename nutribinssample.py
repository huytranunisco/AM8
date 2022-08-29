import time
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import BNPauto

start_time = time.time()

billingAcc = 'nutribin'
accName = 'NutribinsLLC'
facility = 'Joliet'
startPeriod = '08/01/22'
endPeriod = '08/15/22'
user = 'kenguyen'
billCycle = 'Bimonthly'

facility = BNPauto.exportHandle(billingAcc, facility, startPeriod, endPeriod, user)
facility = facility.lower()
facility = facility.capitalize()

reportLoc = BNPauto.invoiceToReport(user, accName, facility, billCycle)
activityLoc = r"C:\\Users\\kenguyen\\Documents\\SOPS\\Nutribins LLC\\Activity Reports\\activityReoport(Nutribins LLC - 2022-08-01 - 2022-08-15).xlsx"

report = pd.read_excel(reportLoc)
billingitems = report['Description'].tolist()
for count, name in enumerate(billingitems):
    billingitems[count] = name.lower()

print(billingitems)

book = load_workbook(reportLoc)
writer = pd.ExcelWriter(reportLoc, engine = 'openpyxl')
writer.book = book

def dataCopy(dataframe, sheetName):
    if sheetName in writer.sheets:
        dataframe.to_excel(writer, sheet_name = sheetName,startrow = writer.sheets[sheetName].max_row, index = False, header = False)
    else:
        dataframe.to_excel(writer, sheet_name = sheetName, index = False)

    writer.save()

# Handling Pick per Pallet
def palletPick(arPath):
    df = pd.read_excel(io=arPath, sheet_name='Pick Task')
    sum = df['PALLETPICKQTY'].sum()

    dataCopy(df, 'PALLET PICK')
    return df, sum

# Handling Pick per Case, WeightRangePerCase,0 - 30;
def pickCase_0_to_30(arPath):
    df = pd.read_excel(io=arPath, sheet_name='Pick Task')
    df2 = df[(df['INDIVIDUAL ITEM WGT'] >= 0) & (df['INDIVIDUAL ITEM WGT'] <= 30)]
    sum = df2['CASEPICKQTY'].sum()

    return df2, sum

# handling order processing per order, ordertype,rg;outbound shipmethod,ltl;
def orderProcessingperOrder_RG_LTL(arPath):
    df = pd.read_excel(io=arPath, sheet_name='Order & Receipt')
    df2 = df[df['Ship Method'].isin(['LTL'])]
    df2 = df2[df2['RN/DN'].str.contains('DN', case=False, na=False)]
    count = df2['RN/DN'].count()

    return df2, count

# handling pick per piece, weightrangepereach,over 30;
def pickPiece_over30(arPath):
    df = pd.read_excel(io=arPath, sheet_name='Pick Task')
    df2 = df[df['INDIVIDUAL ITEM WGT'] > 30]
    sum = df2['PIECEPICKQTY'].sum()

    return df2, sum

# materials & other charges-pallet charge grade b
def palletChargeGradeB(arPath):
    df = pd.read_excel(io=arPath, sheet_name='Materials')
    df2 = df[df['ITEM DESCRIPTION'].str.contains('grade b', case=False, na=False)]
    sum = df2['QTY'].sum()

    return df2, sum

# materials & other charges-stretch wrap
def stretchWrap(arPath):
    df = pd.read_excel(io=arPath, sheet_name='Materials')
    df2 = df[df['ITEM DESCRIPTION'].str.contains('stretch wrap', case=False, na=False)]
    sum = df2['QTY'].sum()

    return df2, sum

# storage income initial storage per pallet, locationtype,1 high;palletsize,none;
def initialStorage(arPath):
    df = pd.read_excel(io=activityLoc, sheet_name='Receive Task')
    sum = df['PALLET QTY'].sum()

    return df, sum

# customized shipping documents
def customizedShippingDocs(arPath, index):
    df = pd.read_excel(io=arPath, sheet_name='Accessories')
    df2 = df[df['DESCRIPTION'].str.contains('CUSTOMIZED SHIPPING DOCUMENTS', case=False, na=False)]
    sum = df2['QTY'].sum()

    return df2, sum

for index, item in enumerate(billingitems):
    print(item)
    if item == 'handling pick per pallet':
        df, qty = palletPick(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'PALLET PICK')
    elif item == 'handling pick per case, weightrangepercase,0 - 30;':
        df, qty = pickCase_0_to_30(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'PICK PER CASE 0 - 30')
    elif item == 'handling order processing per order, ordertype,rg;outbound shipmethod,ltl;':
        df, qty = orderProcessingperOrder_RG_LTL(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'HANDLING ORDER PROCESSING LTL')
    elif item == 'handling pick per piece, weightrangepereach,over 30;':
        df, qty = pickPiece_over30(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'PICK PER PIECE OVER 30')
    elif item == 'materials & other charges-pallet charge grade b' or item == 'accessorial charge grade b pallet':
        df, qty = palletChargeGradeB(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'GRADE B PALLET')
    elif item == 'materials & other charges-stretch wrap' or item == 'accessorial charge stretch wrap':
        df, qty = stretchWrap(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'STRETCH WRAP')
    elif item == 'storage income initial storage per pallet, locationtype,1 high;palletsize,none;':
        df, qty = initialStorage(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'INITIAL STORAGE')
    elif item == 'handling offload per pallet, offloadtype,palletized;mixedmode,singlesku;':
        df, qty = initialStorage(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'HANDLING OFFLOAD PER PALLET')
    elif item == 'customized shipping documents':
        df, qty = customizedShippingDocs(activityLoc)
        report['WISE Qty'][index] = qty
        dataCopy(df, 'CUSTOMIZED SHIPPING DOC')

report.to_excel(reportLoc, index=False)

print("--- %s seconds ---" % (time.time() - start_time))